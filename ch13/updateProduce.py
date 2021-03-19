#! python3
# updateProduce.py - Update costs in produceSales.xlsx

import openpyxl


wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

# produce type and price to update
PRICE_UPDATES = {
    "Garlic": 3.07,
    "Celery": 1.19,
    "Lemon": 1.27}


# loop through the rows and update prices
for rowNum in range(2, sheet.max_row):                          # first row is the header
    produceName = sheet.cell(row = rowNum, column = 1).value    # first column is name of produce
    if produceName in PRICE_UPDATES:                            # if that produce type needs new price
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName] # update price


wb.save("updatedProduceSales.xlsx")                             # save as a new spreadsheet