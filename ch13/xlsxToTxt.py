#! python3
# save spreadsheet to txt file where each row = a line and each column = txt file

import openpyxl

xlsxFile = input("enter excel filename (no extension):")
header = input("is first row a header? (y or n)")

wb = openpyxl.open(xlsxFile + ".xlsx", data_only=True)
sheets = wb.worksheets

print("Reading spreadsheet ...")

# for each worksheet in the file...
for ws in sheets:
    row_count = ws.max_row
    col_count = ws.max_column
    # for each column in the worksheet...
    for c in range(1, col_count + 1):
        # decide what to name the text file
        if header == "y":                   # use header if available
            title = str(ws.cell(row=1, column=c).value)
        else:                               # otherwise use the xlsx name + column number
            title = f"{xlsxFile}_{c}"
        # write file
        with open(f"{title}.txt", "w") as txt:
            for r in range(2, row_count + 1):
                txt.write(str(ws.cell(row=r, column=c).value) + "\n")
        txt.close()

print("Done.")