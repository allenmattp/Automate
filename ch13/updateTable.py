import openpyxl

wb = openpyxl.load_workbook("outstanding.xlsx")

sheets = wb.sheetnames      # create list of sheets in workbook
sheet = wb[sheets[0]]       # set name of first sheet

rows = sheet.max_row        # find how many rows are in sheet
cols = sheet.max_column     # find how many columns are in sheet

department = "name"         # set variable to track department name
for r in range(rows):                                               # iterate through all rows in sheet
    if sheet.cell(row = r+1, column = 1).value:                     # if column 1 has a value
        department = sheet.cell(row = r+1, column = 1).value        # set dept variable to that value
    if sheet.cell(row = r+1, column = 2).value:                     # if column 2 has a value
        sheet.cell(row=r + 1, column=1).value = department          # make sure col 1 has a dept

for r in range(rows, 0, -1):                                        # iterate through all rows backward
    if not sheet.cell(row = r, column = 2).value:                   # if col 2 has no value
        sheet.delete_rows(r, 1)                                     # delete it



wb.save("updatedTable.xlsx")                                    # save as a new spreadsheet