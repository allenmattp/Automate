import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(900))


wb = openpyxl.load_workbook("example.xlsx")

sheets = wb.sheetnames
sheet = wb[sheets[0]]

rows = sheet.max_row
cols = sheet.max_column

print(get_column_letter(cols))
print(column_index_from_string(get_column_letter(cols)))

alphaCols = get_column_letter(cols)

print(sheet)
print(sheet["A1"].value)
print(sheet.cell(row = 1, column = 1).value)

for r in range(rows):
    print(sheet.cell(row = r + 1, column = 1).value)

tuple(sheet["A1":(alphaCols + str(rows))])

for r in sheet["A1":(alphaCols + str(rows))]:
    for cell in r:
        print(cell.coordinate, cell.value)
    print("--- END OF ROW ---")

for c in list(sheet.columns)[1]:
    print(c.value)