#! python3
# inserts blank row(s) in specified file if you don't want to use the built method for some reason

import openpyxl

file = input("enter filename:")
rowLocation = int(input("enter row number to insert first row:"))
rowNumber = int(input("enter number of rows to insert:"))

print("Opening workbook ...")
wb = openpyxl.load_workbook(file)
ws = wb.active

rows = ws.max_row
cols = ws.max_column

print("Inserting rows ...")

for r in range(rows + 1, rowLocation - 1, -1):
    for c in range(1, cols + 1):
        ws.cell(row= r + rowNumber, column=c).value = ws.cell(row=r, column=c).value

for r in range(rowNumber):
    for c in range (1, cols + 1):
        ws.cell(row= rowLocation + r, column=c).value = ""

wb.save(file)

print("Done.")