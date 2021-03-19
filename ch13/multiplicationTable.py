#! python3
# multiplicationTable.py - builds a "number" x "number" multiplication table in excel

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Multiplication Table"

# user specifies table dimensions
number = int(input("Enter a number: "))

print(f"Building a table that is {number} x {number} ...")

# set initial column/row
fontObjLabel = Font(bold = True)
for r in range(2, number+2):
    sheet.cell(row = r, column = 1).value = r - 1
    sheet.cell(row = r, column = 1).font = fontObjLabel
    sheet.cell(row = 1, column = r).value = r - 1
    sheet.cell(row = 1, column = r).font = fontObjLabel

# build the multiplication table
for r in range(2, number+2):
    for c in range(2, number+2):
        sheet.cell(row = r, column = c).value = \
            sheet.cell(row = r, column = 1).value * sheet.cell(row = 1, column = c).value

wb.save("Multiplication Table.xlsx")

print("Done.")