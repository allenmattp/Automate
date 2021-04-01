#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet

import openpyxl, smtplib, passwordGen


# Open spreadsheet and check dues status
wb = openpyxl.load_workbook("duesRecords.xlsx")
ws = wb["Sheet1"]
lastCol = ws.max_column
latestMonth = ws.cell(row=1, column=lastCol).value

# Check each member's payment status
unpaidMembers = {}
for r in range(2, ws.max_row + 1):
    payment = ws.cell(row=r, column=lastCol).value      # find payment status of latest month
    if payment != "paid":                               # check payment status
        name = ws.cell(row=r, column=1).value           # member name stored in first column
        email = ws.cell(row=r, column=2).value          # member email stored in second column
        unpaidMembers[name] = email

# Log into email account
connSmtp = smtplib.SMTP("smtp.gmail.com", 587)
connSmtp.ehlo()
connSmtp.starttls()
connSmtp.login("allenmattpdev@gmail.com", passwordGen.passGen())

# Send out reminder emails
for name, email in unpaidMembers.items():
    body = f"Subject: {latestMonth} dues unpaid.\n" \
           f"Dear {name}:\n" \
           f"Records show that you have not paid dues for {latestMonth}.\n" \
           f"Please make this payment ASAP."
    print(f"Sending email to {email} ...")
    sendStatus = connSmtp.sendmail("allenmattpdev@gmail.com", email, body)

    if sendStatus != {}:
        print(f"There was a problem sending email to {email}: {sendStatus}")

connSmtp.quit()